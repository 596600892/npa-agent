from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


EXPORT_HEADERS = [
    "批次",
    "层级",
    "优先级",
    "脱敏债务人",
    "本金",
    "地区",
    "法院",
    "手机号是否完整",
    "地址是否完整",
    "身份证是否完整",
    "建议动作",
    "合规话术",
    "当前状态",
    "最新跟进结果",
    "下一步",
]


def build_execution_export(tasks: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "执行清单"
    ws.append(EXPORT_HEADERS)
    header_fill = PatternFill("solid", fgColor="E8EFEB")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="123C3A")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for task in tasks:
        ws.append(
            [
                task.get("batch_name"),
                task.get("tier"),
                task.get("priority_score"),
                task.get("masked_debtor"),
                task.get("principal"),
                task.get("region"),
                task.get("court"),
                "是" if task.get("phone_present") else "否",
                "是" if task.get("address_present") else "否",
                "是" if task.get("id_card_present") else "否",
                task.get("suggested_action"),
                task.get("script"),
                task.get("status"),
                task.get("latest_result") or "",
                task.get("next_action"),
            ]
        )
    widths = [18, 8, 10, 14, 14, 12, 24, 14, 14, 14, 18, 48, 14, 18, 28]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
