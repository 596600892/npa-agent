from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]

HEADERS = [
    "项目名称",
    "资产类型",
    "地区",
    "法院",
    "户数",
    "本金总额",
    "成交价",
    "回款金额",
    "回款周期（月）",
    "处置方式",
    "调解成功率",
    "诉讼成功率",
    "执行结果",
    "失败原因",
    "备注",
]

FIELD_GUIDE = [
    ["字段名", "是否必填", "说明"],
    ["项目名称", "否", "历史资产包或处置项目名称"],
    ["资产类型", "否", "个贷、消费贷、房抵、企业贷等"],
    ["地区", "否", "主要处置地区或债务人集中地区"],
    ["法院", "否", "主要管辖或执行法院"],
    ["户数", "否", "历史项目户数"],
    ["本金总额", "条件必填", "本金总额和回款金额至少需要一个"],
    ["成交价", "否", "实际收购价或买包价格"],
    ["回款金额", "条件必填", "本金总额和回款金额至少需要一个"],
    ["回款周期（月）", "否", "从收购到主要回款的周期"],
    ["处置方式", "否", "电话调解、诉讼、执行、分包等"],
    ["调解成功率", "否", "可填 0.18 或 18%"],
    ["诉讼成功率", "否", "可填 0.65 或 65%"],
    ["执行结果", "否", "执行反馈、终本、部分回款等"],
    ["失败原因", "否", "失联、无财产、法院慢、证据不足等"],
    ["备注", "否", "人工经验和补充说明"],
]

TEMPLATE_ROWS = [
    HEADERS,
    ["深圳个贷A包", "个贷", "广东深圳", "深圳市南山区人民法院", 820, 18600000, 930000, 2680000, 14, "电话调解+批量诉讼", 0.22, 0.68, "部分执行回款", "部分失联", "示例行，可删除"],
    ["长沙消费贷B包", "消费贷", "湖南长沙", "长沙市芙蓉区人民法院", 460, 7200000, 288000, 690000, 20, "电话调解", 0.16, 0.42, "调解为主", "触达一般", "示例行，可删除"],
]

SAMPLE_ROWS = [
    HEADERS,
    ["深圳个贷A包", "个贷", "广东深圳", "深圳市南山区人民法院", 820, 18600000, 930000, 2680000, 14, "电话调解+批量诉讼", 0.22, 0.68, "部分执行回款", "部分失联", "深圳法院立案反馈较稳定"],
    ["深圳个贷B包", "消费贷", "广东深圳", "深圳市南山区人民法院", 610, 12100000, 605000, 1620000, 16, "电话调解", 0.2, 0.58, "调解后少量诉讼", "无财产", "中小额户适合先调解"],
    ["广州消费贷A包", "消费贷", "广东广州", "广州市越秀区人民法院", 390, 8200000, 328000, 520000, 28, "分包清收", 0.08, 0.36, "执行周期长", "法院慢", "报价需保守"],
    ["长沙消费贷A包", "消费贷", "湖南长沙", "长沙市芙蓉区人民法院", 470, 7300000, 292000, 780000, 18, "电话调解", 0.17, 0.44, "部分回款", "触达一般", "可作为湖南样本"],
    ["南宁个贷A包", "个贷", "广西南宁", "南宁市青秀区人民法院", 260, 5300000, 180000, 210000, 34, "诉讼执行", 0.05, 0.28, "终本较多", "无财产", "执行预期较弱"],
]


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="123C3A")
    header_font = Font(name="Arial", color="FFFFFF", bold=True)
    body_font = Font(name="Arial", color="17211F")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [18, 12, 14, 24, 10, 14, 14, 14, 14, 18, 14, 14, 18, 18, 28]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A2"


def build_workbook(rows: list[list], include_guide: bool = True) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "历史处置数据"
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    if include_guide:
        guide = wb.create_sheet("字段说明")
        for row in FIELD_GUIDE:
            guide.append(row)
        style_sheet(guide)
        guide.column_dimensions["A"].width = 18
        guide.column_dimensions["B"].width = 14
        guide.column_dimensions["C"].width = 60
    return wb


def main() -> None:
    (ROOT / "templates").mkdir(exist_ok=True)
    (ROOT / "samples").mkdir(exist_ok=True)
    build_workbook(TEMPLATE_ROWS).save(ROOT / "templates" / "公司历史处置数据模板.xlsx")
    build_workbook(SAMPLE_ROWS).save(ROOT / "samples" / "company_history_sample.xlsx")


if __name__ == "__main__":
    main()
