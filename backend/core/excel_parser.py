from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass
class RawSheet:
    sheet_name: str
    headers: list[str]
    rows: list[dict[str, Any]]


def parse_excel(file_path: str | Path, sheet_name: str | None = None) -> RawSheet:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    selected = sheet_name if sheet_name in workbook.sheetnames else workbook.sheetnames[0]
    sheet = workbook[selected]
    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        return RawSheet(selected, [], [])

    header_index = 0
    for index, row in enumerate(all_rows[:10]):
        nonempty = [cell for cell in row if cell not in (None, "")]
        if len(nonempty) >= 2:
            header_index = index
            break

    headers = [str(cell).strip() if cell not in (None, "") else f"未命名列{idx + 1}" for idx, cell in enumerate(all_rows[header_index])]
    rows: list[dict[str, Any]] = []
    for raw in all_rows[header_index + 1 :]:
        if not any(cell not in (None, "") for cell in raw):
            continue
        row = {headers[idx]: raw[idx] if idx < len(raw) else None for idx in range(len(headers))}
        rows.append(row)
    return RawSheet(selected, headers, rows)
