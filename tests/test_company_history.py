from __future__ import annotations

import base64
import json
import shutil
import tempfile
import threading
import unittest
import urllib.request
from pathlib import Path

from openpyxl import Workbook, load_workbook

import backend.app as app
from backend.core.company_history import amount_bucket, build_company_history_analytics, build_court_profiles, normalize_history_rows
from backend.core.company_history_mapping import preview_history_mapping
from backend.core.excel_parser import parse_excel
from backend.core.history_calibrator import build_pricing_calibration
from backend.storage import db

ROOT = Path(__file__).resolve().parents[1]


def make_alias_history_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "历史别名"
    ws.append(["包名", "类型", "城市", "处置法院", "客户数", "本金合计", "买包价格", "回款总额", "处置周期", "清收方式", "和解率", "胜诉率", "未回款原因"])
    ws.append(["深圳别名样本", "消费贷", "广东深圳", "深圳市南山区人民法院", 120, 3000000, 150000, 480000, 13, "电话调解", "18%", "62%", "部分失联"])
    wb.save(path)


class CompanyHistoryCoreTests(unittest.TestCase):
    def test_history_template_can_open(self):
        wb = load_workbook(ROOT / "templates" / "公司历史处置数据模板.xlsx", read_only=True)
        self.assertIn("历史处置数据", wb.sheetnames)
        headers = [cell.value for cell in next(wb["历史处置数据"].iter_rows(max_row=1))]
        self.assertIn("本金总额", headers)
        self.assertIn("回款金额", headers)

    def test_alias_mapping_and_court_profile(self):
        tmpdir = Path(tempfile.mkdtemp(prefix="npa-history-core-"))
        try:
            path = tmpdir / "history.xlsx"
            make_alias_history_workbook(path)
            raw = parse_excel(path)
            preview = preview_history_mapping(raw.headers, raw.rows)
            mapping = {field: item["source_column"] for field, item in preview.items()}
            self.assertEqual(mapping["court_name"], "处置法院")
            self.assertEqual(mapping["purchase_price"], "买包价格")
            self.assertEqual(mapping["recovered_amount"], "回款总额")
            records, errors = normalize_history_rows(raw.rows, mapping, "batch")
            self.assertFalse(errors)
            self.assertAlmostEqual(records[0]["derived"]["recovery_rate"], 0.16)
            self.assertAlmostEqual(records[0]["mediation_success_rate"], 0.18)
            profiles = build_court_profiles(records)
            self.assertEqual(profiles[0]["court_name"], "深圳市南山区人民法院")
            self.assertIn(profiles[0]["label"], {"efficient", "normal", "cautious", "difficult"})
            self.assertEqual(records[0]["derived"]["amount_bucket"], "2w-10w")
            self.assertEqual(profiles[0]["primary_amount_bucket"], "2w-10w")
            self.assertIn("电话调解", profiles[0]["disposal_method_distribution"])
        finally:
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_amount_buckets_and_calibration_explain_match_dimensions(self):
        self.assertEqual(amount_bucket(3000), "0-5k")
        self.assertEqual(amount_bucket(12000), "5k-2w")
        self.assertEqual(amount_bucket(50000), "2w-10w")
        self.assertEqual(amount_bucket(150000), "10w+")
        records = [
            {
                "id": "hist_a",
                "project_name": "深圳同院样本",
                "asset_type": "consumer_loan",
                "region": "广东深圳",
                "court_name": "深圳市南山区人民法院",
                "account_count": 10,
                "principal_total": 100000,
                "recovered_amount": 16000,
                "recovery_months": 12,
                "disposal_method": "电话调解",
                "derived": {"recovery_rate": 0.16, "amount_bucket": "5k-2w"},
            },
            {
                "id": "hist_b",
                "project_name": "弱匹配样本",
                "asset_type": "enterprise_loan",
                "region": "浙江杭州",
                "court_name": "杭州市西湖区人民法院",
                "account_count": 2,
                "principal_total": 400000,
                "recovered_amount": 12000,
                "recovery_months": 30,
                "disposal_method": "诉讼执行",
                "derived": {"recovery_rate": 0.03, "amount_bucket": "10w+"},
            },
        ]
        accounts = [
            {"principal": 9000, "address": "广东省深圳市南山区", "derived": {"id_region": "广东"}, "optional": {"jurisdiction_court": "深圳市南山区人民法院"}},
            {"principal": 11000, "address": "广东省深圳市福田区", "derived": {"id_region": "广东"}, "optional": {"jurisdiction_court": "深圳市南山区人民法院"}},
        ]
        calibration = build_pricing_calibration({"asset_type": "consumer_loan"}, accounts, records, build_court_profiles(records))
        first = calibration["matched_records"][0]
        self.assertEqual(calibration["project_context"]["amount_bucket"], "5k-2w")
        self.assertIn("court", first["matched_dimensions"])
        self.assertIn("region", first["matched_dimensions"])
        self.assertIn("amount_bucket", first["matched_dimensions"])
        self.assertIn("命中同法院", first["match_reason"])
        self.assertIn("by_amount_bucket", calibration["breakdown"])
        self.assertLessEqual(abs(calibration["adjustment"]), 0.03)

    def test_company_history_analytics_groups_records(self):
        records = [
            {"region": "广东深圳", "court_name": "深圳市南山区人民法院", "principal_total": 100000, "recovery_months": 12, "disposal_method": "电话调解", "derived": {"recovery_rate": 0.16, "amount_bucket": "5k-2w"}},
            {"region": "广东深圳", "court_name": "深圳市福田区人民法院", "principal_total": 200000, "recovery_months": 18, "disposal_method": "诉讼执行", "derived": {"recovery_rate": 0.08, "amount_bucket": "2w-10w"}},
        ]
        analytics = build_company_history_analytics(records)
        self.assertEqual(analytics["total_records"], 2)
        self.assertEqual(analytics["usable_recovery_count"], 2)
        self.assertEqual(analytics["by_region"][0]["key"], "广东深圳")
        self.assertTrue(any(item["key"] == "5k-2w" for item in analytics["by_amount_bucket"]))


class CompanyHistoryApiTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp(prefix="npa-history-api-")
        cls.old_db = db.DB_PATH
        cls.old_data = db.DATA_DIR
        db.DB_PATH = Path(cls.tmpdir) / "app.sqlite"
        db.DATA_DIR = Path(cls.tmpdir)
        db.init_db()
        cls.server = app.ThreadingHTTPServer(("127.0.0.1", 0), app.Handler)
        cls.port = cls.server.server_address[1]
        cls.thread = threading.Thread(target=cls.server.serve_forever, daemon=True)
        cls.thread.start()

    @classmethod
    def tearDownClass(cls):
        cls.server.shutdown()
        cls.thread.join(timeout=3)
        cls.server.server_close()
        db.DB_PATH = cls.old_db
        db.DATA_DIR = cls.old_data
        shutil.rmtree(cls.tmpdir, ignore_errors=True)

    def post(self, path, payload):
        data = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(
            f"http://127.0.0.1:{self.port}{path}",
            data=data,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            return json.loads(resp.read().decode("utf-8"))

    def get(self, path):
        with urllib.request.urlopen(f"http://127.0.0.1:{self.port}{path}", timeout=10) as resp:
            return json.loads(resp.read().decode("utf-8"))

    def test_history_import_profiles_and_pricing_calibration_report(self):
        history_content = base64.b64encode((ROOT / "samples" / "company_history_sample.xlsx").read_bytes()).decode("ascii")
        uploaded = self.post("/api/company-history/files", {"filename": "company_history_sample.xlsx", "content_base64": history_content})
        preview = self.post("/api/company-history/field-mapping/preview", {"file_id": uploaded["file"]["id"]})
        self.assertEqual(preview["mapping"]["court_name"]["source_column"], "法院")
        mapping = {field: item["source_column"] for field, item in preview["mapping"].items()}
        confirmed = self.post("/api/company-history/field-mapping/confirm", {"file_id": uploaded["file"]["id"], "mapping": mapping, "confidence": preview["mapping"]})
        self.assertEqual(confirmed["normalized_count"], 5)
        self.assertGreaterEqual(confirmed["court_profile_count"], 3)
        analytics = self.get("/api/company-history/analytics")["analytics"]
        self.assertGreaterEqual(analytics["total_records"], 5)
        self.assertTrue(analytics["by_amount_bucket"])
        profiles = self.get("/api/courts/profiles")["profiles"]
        self.assertTrue(any(item["court_name"] == "深圳市南山区人民法院" for item in profiles))
        self.assertTrue(any(item.get("primary_amount_bucket") for item in profiles))

        project = self.post("/api/projects", {"name": "历史校准 API 样例", "asset_type": "consumer_loan"})["project"]
        sample_content = base64.b64encode((ROOT / "samples" / "level3_court.xlsx").read_bytes()).decode("ascii")
        asset_file = self.post(f"/api/projects/{project['id']}/files", {"filename": "level3_court.xlsx", "file_type": "asset_package_excel", "content_base64": sample_content})
        asset_preview = self.post(f"/api/projects/{project['id']}/field-mapping/preview", {"file_id": asset_file["file"]["id"]})
        asset_mapping = {field: item["source_column"] for field, item in asset_preview["mapping"].items()}
        self.post(f"/api/projects/{project['id']}/field-mapping/confirm", {"file_id": asset_file["file"]["id"], "mapping": asset_mapping, "confidence": asset_preview["mapping"]})
        analysis = self.post(f"/api/projects/{project['id']}/analysis/run", {"analysis_type": "consumer_loan_initial_screening"})
        self.assertTrue(analysis["ok"])
        report = self.get(f"/api/projects/{project['id']}/reports/latest")["report"]
        self.assertIn("公司历史校准", report["markdown"])
        self.assertIn("法院画像", report["markdown"])
        self.assertIn("匹配维度解释", report["markdown"])
        self.assertIn("样本可信度", report["markdown"])
        self.assertIn("分维度对比", report["markdown"])
        self.assertGreater(report["data"]["calibration"]["matched_count"], 0)
        latest = self.get(f"/api/projects/{project['id']}/calibration/latest")["calibration"]
        self.assertEqual(latest["project_id"], project["id"])
        self.assertIn("breakdown", latest["calibration"])
        self.assertIn("sample_confidence", latest["calibration"])


if __name__ == "__main__":
    unittest.main()
