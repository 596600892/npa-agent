from __future__ import annotations

import base64
import json
import shutil
import tempfile
import threading
import unittest
import urllib.request
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

import backend.app as app
from backend.storage import db

ROOT = Path(__file__).resolve().parents[1]


class ExecutionApiTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp(prefix="npa-execution-api-")
        cls.old_db = db.DB_PATH
        cls.old_data = db.DATA_DIR
        db.DB_PATH = Path(cls.tmpdir) / "app.sqlite"
        db.DATA_DIR = Path(cls.tmpdir)
        db.init_db()
        cls.server = app.ThreadingHTTPServer(("127.0.0.1", 0), app.Handler)
        cls.port = cls.server.server_address[1]
        cls.thread = threading.Thread(target=cls.server.serve_forever, daemon=True)
        cls.thread.start()

    @classmethod
    def tearDownClass(cls):
        cls.server.shutdown()
        cls.thread.join(timeout=3)
        cls.server.server_close()
        db.DB_PATH = cls.old_db
        db.DATA_DIR = cls.old_data
        shutil.rmtree(cls.tmpdir, ignore_errors=True)

    def request(self, method, path, payload=None):
        data = json.dumps(payload or {}).encode("utf-8") if payload is not None else None
        req = urllib.request.Request(
            f"http://127.0.0.1:{self.port}{path}",
            data=data,
            headers={"Content-Type": "application/json"},
            method=method,
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            body = resp.read()
            if resp.headers.get_content_type() == "application/json":
                return json.loads(body.decode("utf-8"))
            return body

    def post(self, path, payload):
        return self.request("POST", path, payload)

    def get(self, path):
        return self.request("GET", path)

    def patch(self, path, payload):
        return self.request("PATCH", path, payload)

    def create_analyzed_project(self):
        project = self.post("/api/projects", {"name": "执行 API 样例", "asset_type": "consumer_loan"})["project"]
        content = base64.b64encode((ROOT / "samples" / "level3_court.xlsx").read_bytes()).decode("ascii")
        uploaded = self.post(f"/api/projects/{project['id']}/files", {"filename": "level3_court.xlsx", "file_type": "asset_package_excel", "content_base64": content})
        preview = self.post(f"/api/projects/{project['id']}/field-mapping/preview", {"file_id": uploaded["file"]["id"]})
        mapping = {field: item["source_column"] for field, item in preview["mapping"].items()}
        self.post(f"/api/projects/{project['id']}/field-mapping/confirm", {"file_id": uploaded["file"]["id"], "mapping": mapping, "confidence": preview["mapping"]})
        self.post(f"/api/projects/{project['id']}/analysis/run", {"analysis_type": "consumer_loan_initial_screening"})
        return project

    def test_execution_plan_tasks_update_and_event_flow(self):
        project = self.create_analyzed_project()
        generated = self.post(f"/api/projects/{project['id']}/execution/plan", {})
        self.assertTrue(generated["ok"])
        self.assertEqual(generated["summary"]["task_count"], 12)
        self.assertGreaterEqual(len(generated["batches"]), 5)
        task = generated["tasks"][0]
        self.assertIn(task["status"], {"pending"})
        batches = self.get(f"/api/projects/{project['id']}/execution/batches")
        self.assertEqual(batches["plan"]["id"], generated["plan"]["id"])
        tasks = self.get(f"/api/projects/{project['id']}/execution/tasks")
        self.assertEqual(len(tasks["tasks"]), 12)
        patched = self.patch(
            f"/api/projects/{project['id']}/execution/tasks/{task['id']}",
            {"status": "contacted", "latest_note": "已接通，核实身份。", "next_action": "登记协商意愿"},
        )
        self.assertEqual(patched["task"]["status"], "contacted")
        event = self.post(
            f"/api/projects/{project['id']}/execution/tasks/{task['id']}/events",
            {"event_type": "contact_result", "result": "willing", "note": "愿意协商分期", "next_action": "生成分期协商方案"},
        )
        self.assertEqual(event["task"]["status"], "willing")
        self.assertEqual(event["task"]["latest_result"], "willing")
        filtered = self.get(f"/api/projects/{project['id']}/execution/tasks?status=willing")
        self.assertEqual(len(filtered["tasks"]), 1)
        exported = self.request("GET", f"/api/projects/{project['id']}/execution/export.xlsx")
        wb = load_workbook(BytesIO(exported), read_only=True)
        ws = wb["执行清单"]
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        self.assertIn("批次", headers)
        self.assertIn("合规话术", headers)
        text = "\n".join(str(cell.value or "") for row in ws.iter_rows() for cell in row)
        self.assertNotIn("440305199001011234", text)
        self.assertNotIn("13812345678", text)
        self.assertNotIn("广东省深圳市南山区科技园", text)
        rerun = self.post(f"/api/projects/{project['id']}/analysis/run", {"analysis_type": "consumer_loan_initial_screening"})
        self.assertIn("处置执行计划", rerun["report"]["markdown"])
        self.assertIn("已生成执行任务", rerun["report"]["markdown"])


if __name__ == "__main__":
    unittest.main()
